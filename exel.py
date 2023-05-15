# import xlsxwriter module
import xlsxwriter
import openpyxl as xl
from pathlib import Path
from os.path import exists

# creates base file if it doesnt exist

# -----------------------------------------------------------------
# Workbook() takes one, non-optional, argument
# which is the filename that we want to create.
p = Path(__file__).with_name('japhy.xlsx')
if (not exists(p)):
    
    workbook = xlsxwriter.Workbook('japhy.xlsx')

    # The workbook object is then used to add new
    # worksheet via the add_worksheet() method.
    worksheet = workbook.add_worksheet()

    # Use the worksheet object to write
    # data via the write() method.
    worksheet.write('A1', 'sex')
    worksheet.write('B1', 'castrer')
    worksheet.write('C1', 'race')
    worksheet.write('D1', 'age')
    worksheet.write('E1', 'activity')
    worksheet.write('F1', 'fitness')
    worksheet.write('G1', 'weight')
    worksheet.write('H1', 'menu item 1')
    worksheet.write('I1', 'menu item 2')


    # Finally, close the Excel file
    # via the close() method.
    workbook.close()


# -----------------------------------------------------------------

def add_data(array):
    row  = firstEmptyRow()
    
    p = Path(__file__).with_name('japhy.xlsx')
    book = xl.load_workbook(p)
    ws = book.worksheets[0]
    

    ws.cell(row, 1, array[0])
    ws.cell(row, 2, array[1])
    ws.cell(row, 3, array[2])
    ws.cell(row, 4, array[3])
    ws.cell(row, 5, array[4])
    ws.cell(row, 6, array[5])
    ws.cell(row, 7, array[6])
    ws.cell(row, 8, array[7])
    ws.cell(row, 9, array[8])
    book.save('japhy.xlsx')
    book.close()

# This function returns the index of the first empty row
def firstEmptyRow():
    p = Path(__file__).with_name('japhy.xlsx')
    book = xl.load_workbook(p)
    ws = book.worksheets[0]
    max = ws.max_row + 1
    book.close()
    return max
    